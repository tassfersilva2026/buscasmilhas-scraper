# maxmilhas_scraper_gha.py — GitHub Actions (1 ciclo, headless, 2ª página)
# Saída: data/MAX_YYYYMMDD_HHMMSS.xlsx (aba "MAX")
# Execução local: python maxmilhas_scraper_gha.py --once --headless
# Baseado no loop do anexo (XPaths/colunas e extrações). 2ª página com "Comprar".  :contentReference[oaicite:1]{index=1}

import os, re, time, argparse, logging
from datetime import datetime, timedelta, date, time as dtime
from decimal import Decimal, InvalidOperation
from zoneinfo import ZoneInfo
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

# ===================== CONFIG =====================
TZ = ZoneInfo("America/Sao_Paulo")
SHEET_NAME = "MAX"
CENTER = Alignment(horizontal="center", vertical="center")

ADVP_LIST = [1, 3, 7, 14, 21, 30, 60, 90]
TRECHOS = [
    ("CGH", "SDU"), ("SDU", "CGH"),
    ("GRU", "POA"), ("POA", "GRU"),
    ("CGH", "GIG"), ("GIG", "CGH"),
    ("BSB", "CGH"), ("CGH", "BSB"),
    ("CGH", "REC"), ("REC", "CGH"),
    ("CGH", "SSA"), ("SSA", "CGH"),
    ("BSB", "GIG"), ("GIG", "BSB"),
    ("GIG", "REC"), ("REC", "GIG"),
    ("GIG", "SSA"), ("SSA", "GIG"),
    ("BSB", "SDU"), ("SDU", "BSB"),
]

def build_url(origin: str, destiny: str, departure_date: str) -> str:
    # OW / 1 adulto / Econ
    return f"https://www.maxmilhas.com.br/busca-passagens-aereas/OW/{origin}/{destiny}/{departure_date}/1/0/0/EC"

# ===== XPaths (2ª página) — iguais ao anexo =====
XP_BUY       = '//*[@id="__next"]/div[4]/section/div[5]/div[2]/div[1]/div/div/div[1]/div/div/div[2]/div[2]/div/div[2]/button'
XP_BUY_FALL  = "//button[contains(., 'Comprar') or contains(., 'COMPRAR')]"

XP_HR_PART   = '//*[@id="__next"]/div[4]/section/div[2]/div/div/div/article/div/div[1]/div[1]/div[2]'
XP_HR_CHEG   = '//*[@id="__next"]/div[4]/section/div[2]/div/div/div/article/div/div[1]/div[1]/div[4]'

XP_COL_G_TARIFA = '//*[@id="__next"]/div[4]/section/div[3]/div[1]/div/div[3]/div/div[1]/div/span[2]'
XP_COL_J_TOTAL  = ('//*[@id="__next"]/div[4]/section/div[3]/div[1]/div/div[3]/div/div[4]/span[2]'
                   ' | //*[@id="__next"]/div[3]/div[1]/div/div[3]/div/div[4]/span[2]')
XP_COL_N_TIPO   = ('//*[@id="__next"]/div[4]/section/div[2]/div/div/div/article/div/div[1]/div[2]/ul/li[4]'
                   ' | //*[@id="__next"]//article//div[1]/div[2]//li[contains(., "Tarifa")]')

XP_TX_EMB   = '//*[@id="__next"]/div[3]/div[1]/div/div[3]/div/div[3]/span/div[1]/span[2] | //*[@id="__next"]/div[4]/section/div[3]/div[1]/div/div[3]/div/div[3]/span/div[1]/span[2]'
XP_TX_EMIS  = '//*[@id="__next"]/div[3]/div[1]/div/div[3]/div/div[3]/span/div[2]/span[2] | //*[@id="__next"]/div[4]/section/div[3]/div[1]/div/div[3]/div/div[3]/span/div[2]/span[2]'
XP_DESC     = '//*[@id="__next"]/div[3]/div[1]/div/div[3]/div/div[2]/span/div/span[2] | //*[@id="__next"]/div[4]/section/div[3]/div[1]/div/div[3]/div/div[2]/span/div/span[2]'
XP_CIA      = '//*[@id="__next"]/div[4]/section/div[2]/div/div/div/article/div/div[2]/div/div[1]/ul/li[3]'

HEADERS = [
    "DATA DA BUSCA","HORA DA BUSCA","TRECHO",
    "DATA DO VOO","HR IDA","HR VOLTA",
    "TARIFA","DESCONTO","TX DE EMBARQUE","VALOR COM TAXA",
    "TX DE  EMISSÃO","TOTAL","CIA DO VOO","TIPO (A/C)"
]

# ===================== Helpers =====================
def brl_to_decimal(txt: str):
    if not txt: return None
    s = re.sub(r"[^\d,\.]", "", txt).replace(".", "").replace(",", ".")
    try: return Decimal(s)
    except (InvalidOperation, TypeError): return None

def parse_time_only(txt: str):
    if not txt: return None
    m = re.search(r'(\d{2}):(\d{2})(?::(\d{2}))?', txt.strip())
    if not m: return None
    hh, mm, ss = m.groups(); ss = ss or "00"
    try: return dtime(int(hh), int(mm), int(ss))
    except ValueError: return None

def clean_cia_text(txt: str | None) -> str:
    if not txt: return ""
    s = txt.strip()
    s = re.sub(r'(?i)\blinhas a[eé]reas\b', '', s)
    s = re.sub(r'\s{2,}', ' ', s).strip(" -–,.;:/\t\n\r")
    return s

def extract_letra_tarifa(txt: str | None) -> str | None:
    if not txt: return None
    t = re.sub(r'\s+', ' ', txt).strip()
    m = re.search(r'(?i)tarifa\s*([A-Z])\b', t)
    if m: return m.group(1).upper()
    m2 = re.findall(r'\b([A-Z])\b', t)
    return m2[-1].upper() if m2 else None

def to_excel_naive(v):
    if isinstance(v, datetime): return v.replace(tzinfo=None)
    if isinstance(v, dtime):    return v.replace(tzinfo=None)
    return v

# ===================== Excel =====================
def _is_valid_xlsx(path: str) -> bool:
    try:
        if not os.path.isfile(path): return False
        with ZipFile(path, "r") as zf:
            return "[Content_Types].xml" in zf.namelist()
    except Exception:
        return False

def _create_new_workbook(path: str):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = SHEET_NAME
    ws.append(HEADERS)
    widths = {
        "DATA DA BUSCA":14,"HORA DA BUSCA":12,"TRECHO":12,
        "DATA DO VOO":14,"HR IDA":12,"HR VOLTA":12,
        "TARIFA":14,"DESCONTO":14,"TX DE EMBARQUE":16,"VALOR COM TAXA":16,
        "TX DE  EMISSÃO":16,"TOTAL":14,"CIA DO VOO":22,"TIPO (A/C)":12
    }
    for j, hdr in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(hdr, 16)
        ws.cell(row=1, column=j).alignment = CENTER
    wb.save(path); wb.close()

def ensure_workbook(path: str):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    if not os.path.exists(path) or not _is_valid_xlsx(path):
        _create_new_workbook(path); return
    wb = load_workbook(path)
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(HEADERS)
        for j in range(1, len(HEADERS)+1):
            ws.cell(row=1, column=j).alignment = CENTER
        wb.save(path)
    wb.close()

def append_row(path: str, row_values: list):
    ensure_workbook(path)
    wb = load_workbook(path); ws = wb[SHEET_NAME]
    ws.append([to_excel_naive(v) for v in row_values])
    r = ws.max_row
    numfmt = {
        1:"DD/MM/YYYY", 2:"HH:MM:SS", 4:"DD/MM/YYYY", 5:"HH:MM:SS", 6:"HH:MM:SS",
        7:"#,##0.00", 8:"#,##0.00", 9:"#,##0.00", 10:"#,##0.00", 11:"#,##0.00", 12:"#,##0.00"
    }
    for c in range(1, len(HEADERS)+1):
        cell = ws.cell(row=r, column=c)
        if c in numfmt and cell.value is not None:
            cell.number_format = numfmt[c]
        cell.alignment = CENTER
    wb.save(path); wb.close()

# ===================== Selenium =====================
def _maybe_set_binary_location(opts: Options):
    chrome_path = os.getenv("CHROME_PATH") or os.getenv("GOOGLE_CHROME_SHIM")
    if chrome_path and os.path.exists(chrome_path):
        opts.binary_location = chrome_path

def setup_driver(headless=True):
    options = Options()
    if headless:
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--hide-scrollbars")
    _maybe_set_binary_location(options)
    options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-sandbox")
    options.add_argument("--log-level=3")
    options.add_argument("--mute-audio")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--no-first-run")
    options.add_argument("--disable-notifications")
    options.add_argument("--lang=pt-BR")
    options.add_argument("--use-gl=swiftshader")
    options.add_argument("--enable-unsafe-swiftshader")
    service = ChromeService()
    driver = webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 15)
    return driver, wait

def navigate_same_tab(driver, url):
    try:
        driver.execute_script("window.location.assign(arguments[0]);", url)
    except Exception:
        driver.get(url)

def wait_buy_or_empty(driver, max_wait=20):
    t0 = time.time()
    while time.time() - t0 < max_wait:
        try:
            if WebDriverWait(driver, 1.5).until(EC.presence_of_element_located((By.XPATH, XP_BUY))):
                return "buy"
        except TimeoutException:
            pass
        try:
            btns = [b for b in driver.find_elements(By.XPATH, XP_BUY_FALL) if b.is_displayed()]
            if btns: return "buy"
        except Exception:
            pass
        # “sem resultados”
        try:
            if driver.find_elements(By.XPATH, "//*[contains(., 'Nenhum voo') or contains(., 'sem resultados') or contains(., 'não encontramos')]"):
                return "empty"
        except Exception:
            pass
        time.sleep(1)
    return "timeout"

def click_buy_js(driver) -> bool:
    try:
        el = WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.XPATH, XP_BUY)))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        time.sleep(0.05)
        driver.execute_script("arguments[0].click();", el)
        return True
    except Exception:
        try:
            btns = [b for b in driver.find_elements(By.XPATH, XP_BUY_FALL) if b.is_displayed()]
            btns.sort(key=lambda e: (e.rect.get('y', 1e9), e.rect.get('x', 1e9)))
            for b in btns:
                try:
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", b)
                    time.sleep(0.05)
                    driver.execute_script("arguments[0].click();", b)
                    return True
                except Exception:
                    continue
        except Exception:
            return False
    return False

def get_text(driver, xpath, timeout=10):
    try:
        el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, xpath)))
        return (el.text or "").strip()
    except TimeoutException:
        return None

def get_text_multi(driver, xps, timeout=10):
    for xp in xps:
        txt = get_text(driver, xp, timeout)
        if txt: return txt
    return None

# ===================== 1 busca =====================
def run_one_search(driver, origin, destiny, advp, out_path, espera):
    now = datetime.now(TZ)
    data_busca = now.date()
    hora_busca = dtime(now.hour, now.minute, now.second)
    data_voo   = (now + timedelta(days=advp)).date()
    trecho_str = f"{origin}-{destiny}"
    url = build_url(origin, destiny, data_voo.strftime("%Y-%m-%d"))

    logging.info("Trecho %s | ADVP %d | URL: %s", trecho_str, advp, url)
    navigate_same_tab(driver, url)

    status = wait_buy_or_empty(driver, max_wait=espera)
    if status != "buy":
        values = [data_busca, hora_busca, trecho_str, data_voo, None, None, None, None, None, None, None, None, "Sem Ofertas", ""]
        append_row(out_path, values); return

    old = set(driver.window_handles)
    if not click_buy_js(driver):
        values = [data_busca, hora_busca, trecho_str, data_voo, None, None, None, None, None, None, None, None, "Sem Ofertas", ""]
        append_row(out_path, values); return

    # troca de aba se abrir
    switched = False
    try:
        WebDriverWait(driver, 6).until(lambda d: len(d.window_handles) > len(old))
        newh = [h for h in driver.window_handles if h not in old]
        if newh:
            driver.switch_to.window(newh[-1]); switched = True
    except TimeoutException:
        pass

    # coleta 2ª página
    hr_ida_txt   = get_text(driver, XP_HR_PART, timeout=12)
    hr_volta_txt = get_text(driver, XP_HR_CHEG, timeout=12)

    tarifa_g_txt = get_text(driver, XP_COL_G_TARIFA, timeout=12)
    total_j_txt  = get_text_multi(driver, XP_COL_J_TOTAL.split("|"), timeout=12)
    tx_emb_txt   = get_text_multi(driver, XP_TX_EMB.split("|"),      timeout=12)
    tx_emis_txt  = get_text_multi(driver, XP_TX_EMIS.split("|"),     timeout=12)
    desc_txt     = get_text_multi(driver, XP_DESC.split("|"),        timeout=10)
    cia_txt      = get_text(driver, XP_CIA,                          timeout=10)

    tipo_txt   = get_text_multi(driver, XP_COL_N_TIPO.split("|"), timeout=10)
    tipo_letra = extract_letra_tarifa(tipo_txt) or ""

    hr_ida   = parse_time_only(hr_ida_txt)
    hr_volta = parse_time_only(hr_volta_txt)
    tarifa_g = brl_to_decimal(tarifa_g_txt)
    desconto = brl_to_decimal(desc_txt) if desc_txt else None
    tx_emb   = brl_to_decimal(tx_emb_txt)
    total_j  = brl_to_decimal(total_j_txt)
    tx_emis  = brl_to_decimal(tx_emis_txt)
    total_l  = total_j  # TOTAL = "valor com taxa"

    cia      = clean_cia_text(cia_txt)

    values = [
        data_busca, hora_busca, trecho_str,
        data_voo, hr_ida, hr_volta,
        float(tarifa_g) if tarifa_g is not None else None,
        float(desconto) if desconto is not None else None,
        float(tx_emb) if tx_emb is not None else None,
        float(total_j) if total_j is not None else None,
        float(tx_emis) if tx_emis is not None else None,
        float(total_l) if total_l is not None else None,
        cia or "", tipo_letra
    ]
    append_row(out_path, values)

    if switched:
        try: driver.close()
        except: pass
        try: driver.switch_to.window(driver.window_handles[0])
        except: pass

# ===================== MAIN =====================
def main():
    ap = argparse.ArgumentParser(description="MaxMilhas scraper (CI/Actions).")
    ap.add_argument("--saida",  default="data", help="Pasta para salvar Excel (raiz/data)")
    ap.add_argument("--espera", type=int, default=20, help="Segundos p/ achar 'Comprar'/detalhes")
    ap.add_argument("--headless", action="store_true")
    ap.add_argument("--gui", dest="headless", action="store_false")
    ap.add_argument("--once", action="store_true", help="Roda 1 ciclo e finaliza")
    ap.set_defaults(headless=True)
    args = ap.parse_args()

    os.makedirs(args.saida, exist_ok=True)
    # Nome: MAX_YYYYMMDD_HHMMSS.xlsx (America/Sao_Paulo)
    stamp = datetime.now(TZ).strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(args.saida, f"MAX_{stamp}.xlsx")

    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s | %(levelname)s | %(message)s",
                        datefmt="%H:%M:%S")
    logging.info("Planilha: %s | Aba: %s | Headless: %s", out_path, SHEET_NAME, args.headless)

    ensure_workbook(out_path)
    driver, _wait = setup_driver(headless=args.headless)
    base_tab = driver.current_window_handle

    try:
        def ciclo():
            nonlocal base_tab
            for (origin, destiny) in TRECHOS:
                for advp in ADVP_LIST:
                    try:
                        run_one_search(driver, origin, destiny, advp, out_path=out_path, espera=args.espera)
                    except Exception as e:
                        logging.exception("Falha em %s-%s ADVP %d: %s", origin, destiny, advp, e)

        if args.once:
            ciclo()
        else:
            ciclo()  # no CI normalmente rodamos --once; aqui deixo 1 ciclo por padrão

    finally:
        try: driver.quit()
        except Exception: pass
        logging.info("Finalizado.")

if __name__ == "__main__":
    main()
