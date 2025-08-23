# flipmilhas_scraper_gha.py — versão p/ GitHub Actions (1 ciclo + headless)
# Saída: output/FLIPMILHAS.xlsx (aba "BUSCAS")
# Execução típica (local): python flipmilhas_scraper_gha.py --headless --once
# Execução no GitHub Actions: ver .github/workflows/scrape.yml

import os, re, time, argparse, logging, shutil, tempfile
from datetime import datetime, timedelta, date, time as dtime
from decimal import Decimal, InvalidOperation
from zoneinfo import ZoneInfo
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

# ======= CONFIG PADRÃO =======
TZ = ZoneInfo("America/Sao_Paulo")
SHEET_NAME = "BUSCAS"

ADVP_LIST = [1, 3, 7, 14, 21, 30, 60, 90]
TRECHOS = [
    ("CGH", "SDU"), ("SDU", "CGH"),
    ("GRU", "POA"), ("POA", "GRU"),
    ("CGH", "GIG"), ("GIG", "CGH"),
    ("BSB", "CGH"), ("CGH", "BSB"),
    ("CGH", "REC"), ("REC", "CGH"),
    ("CGH", "SSA"), ("SSA", "CGH"),
    ("BSB", "GIG"), ("GIG", "BSB"),
    ("GIG", "REC"), ("REC", "GIG"),
    ("GIG", "SSA"), ("SSA", "GIG"),
    ("BSB", "SDU"), ("SDU", "BSB"),
]

# XPaths (ajuste se o site mudar)
XPATH_NO_FLIGHTS    = '//*[@id="flights"]/div/h3'
XPATH_TEXTO_COMPRAR = "//button[normalize-space()='Comprar' or contains(translate(.,'COMPRAR','comprar'),'comprar')]"
XPATH_BOTAO_COMPRAR = '/html/body/div[1]/main/section/div[2]/div[3]/div[2]/div/div[1]/div[2]/button'
XPATH_PARTIDA = '//*[@id="app"]/main/section/section/section/section[1]/div[2]/div[1]/p[1]'
XPATH_CHEGADA = '//*[@id="app"]/main/section/section/section/section[1]/div[2]/div[2]/p[1]'
XPATH_TARIFA  = '//*[@id="app"]/main/section/section/section/section[2]/section[2]/section[2]/p[2]'
XPATH_TAXA    = '//*[@id="app"]/main/section/section/section/section[2]/section[2]/section[4]/p[2]'
XPATH_CIA     = '//*[@id="app"]/main/section/section/section/section[1]/div[3]/div[2]/p[2]'

CENTER = Alignment(horizontal="center", vertical="center")
HEADERS = [
    "DATA DA BUSCA","HORA DA BUSCA","TRECHO",
    "DATA PARTIDA","HORA DA PARTIDA","DATA CHEGADA","HORA DA CHEGADA",
    "TARIFA","TX DE EMBARQUE","TOTAL","CIA DO VOO",
]

# ======= Helpers =======
def build_url(origin: str, destiny: str, departure_date: str) -> str:
    return ("https://flipmilhas.com/passagens"
            f"?adults=1&babies=0&back_date=&children=0&class=economica"
            f"&departure_date={departure_date}&destiny={destiny}&origin={origin}&rooms=1")

def brl_to_decimal(txt: str):
    if not txt: return None
    s = re.sub(r"[^\d,\.]", "", txt).replace(".", "").replace(",", ".")
    try: return Decimal(s)
    except (InvalidOperation, TypeError): return None

def parse_datetime_br(txt: str, fallback_date: date | None = None):
    if not txt: return None
    txt = txt.strip()
    m = re.search(r'(\d{2})/(\d{2})/(\d{4})\s+(\d{2}):(\d{2})(?::(\d{2}))?', txt)
    if m:
        d, mth, y, hh, mm, ss = m.groups(); ss = ss or "00"
        try: return datetime(int(y), int(mth), int(d), int(hh), int(mm), int(ss))
        except ValueError: return None
    m = re.search(r'(\d{2})/(\d{2})\s+(\d{2}):(\d{2})(?::(\d{2}))?', txt)
    if m and fallback_date:
        d, mth, hh, mm, ss = m.groups(); ss = ss or "00"
        try: return datetime(fallback_date.year, int(mth), int(d), int(hh), int(mm), int(ss))
        except ValueError: return None
    m = re.search(r'(\d{2}):(\d{2})(?::(\d{2}))?', txt)
    if m and fallback_date:
        hh, mm, ss = m.groups(); ss = ss or "00"
        try: return datetime(fallback_date.year, fallback_date.month, fallback_date.day, int(hh), int(mm), int(ss))
        except ValueError: return None
    return None

def clean_cia_text(txt: str | None) -> str:
    if not txt: return ""
    s = txt.strip()
    s = re.sub(r'(?i)\blinhas a[eé]reas\b(?:\s+(s\.?\s*/?\s*a\.?)\b)?', '', s)
    s = re.sub(r'\s{2,}', ' ', s).strip(" -–,.;:/\t\n\r")
    return s

def to_excel_naive(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, dtime):
        return value.replace(tzinfo=None)
    return value

# ======= Excel =======
def _is_valid_xlsx(path: str) -> bool:
    try:
        if not os.path.isfile(path): return False
        with ZipFile(path, "r") as zf:
            return "[Content_Types].xml" in zf.namelist()
    except Exception:
        return False

def _create_new_workbook(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    widths = {
        "DATA DA BUSCA": 14, "HORA DA BUSCA": 12, "TRECHO": 12,
        "DATA PARTIDA": 14, "HORA DA PARTIDA": 14,
        "DATA CHEGADA": 14, "HORA DA CHEGADA": 14,
        "TARIFA": 14, "TX DE EMBARQUE": 16, "TOTAL": 14, "CIA DO VOO": 20,
    }
    for j, hdr in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(hdr, 16)
        ws.cell(row=1, column=j).alignment = CENTER
    wb.save(path); wb.close()

def ensure_workbook(path: str):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    if not os.path.exists(path) or not _is_valid_xlsx(path):
        _create_new_workbook(path); return
    wb = load_workbook(path)
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(HEADERS)
        for j in range(1, len(HEADERS)+1):
            ws.cell(row=1, column=j).alignment = CENTER
        wb.save(path)
    wb.close()

def append_row(path: str, row_values: dict):
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    if ws.max_row == 1 and all((c.value is None for c in ws[1])):
        ws.append(HEADERS)
        for j in range(1, len(HEADERS)+1):
            ws.cell(row=1, column=j).alignment = CENTER

    clean = {k: to_excel_naive(v) for k, v in row_values.items()}
    row = [clean.get(h) for h in HEADERS]
    ws.append(row)
    r = ws.max_row

    fmt = {
        "DATA DA BUSCA": "DD/MM/YYYY",
        "HORA DA BUSCA": "HH:MM:SS",
        "DATA PARTIDA": "DD/MM/YYYY",
        "HORA DA PARTIDA": "HH:MM:SS",
        "DATA CHEGADA": "DD/MM/YYYY",
        "HORA DA CHEGADA": "HH:MM:SS",
        "TARIFA": "#,##0.00",
        "TX DE EMBARQUE": "#,##0.00",
        "TOTAL": "#,##0.00",
    }
    for c_idx, hdr in enumerate(HEADERS, start=1):
        cell = ws.cell(row=r, column=c_idx)
        if hdr in fmt and cell.value is not None:
            cell.number_format = fmt[hdr]
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(path)
    wb.close()

# ======= Selenium =======
def setup_driver(headless=True):
    options = Options()
    if headless:
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--hide-scrollbars")
    options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-sandbox")
    options.add_argument("--log-level=3")
    options.add_argument("--mute-audio")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--no-first-run")
    options.add_argument("--disable-notifications")
    driver = webdriver.Chrome(options=options)  # Selenium Manager resolve driver/browser
    wait = WebDriverWait(driver, 15)
    return driver, wait

def navigate_same_tab(driver, url):
    try:
        driver.execute_script("window.location.assign(arguments[0]);", url)
    except Exception:
        driver.get(url)

def wait_for_buy_button_or_no_flights(driver, max_wait):
    start = time.time()
    while time.time() - start < max_wait:
        try:
            el = WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, XPATH_NO_FLIGHTS)))
            txt = (el.text or "").strip().lower()
            if "nenhum voo" in txt:
                return "no_flights"
        except TimeoutException:
            pass
        for xp in (XPATH_TEXTO_COMPRAR, XPATH_BOTAO_COMPRAR):
            try:
                if WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, xp))):
                    return "buy_ready"
            except TimeoutException:
                pass
        time.sleep(1)
    return "timeout"

def js_click_first_buy(driver) -> bool:
    try:
        btns = driver.find_elements(By.XPATH, XPATH_TEXTO_COMPRAR)
        btns = [b for b in btns if b.is_displayed()]
        btns.sort(key=lambda e: (e.rect.get('y', 1e9), e.rect.get('x', 1e9)))
        for el in btns:
            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                time.sleep(0.05)
                driver.execute_script("arguments[0].click();", el)
                return True
            except Exception:
                continue
    except Exception:
        pass
    try:
        el = WebDriverWait(driver, 6).until(EC.visibility_of_element_located((By.XPATH, XPATH_BOTAO_COMPRAR)))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        driver.execute_script("arguments[0].click();", el)
        return True
    except TimeoutException:
        return False

def wait_text(driver, xpath, timeout=12):
    try:
        el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, xpath)))
        return (el.text or "").strip()
    except TimeoutException:
        return None

def wait_text_retry(driver, xpath, tries=3, delay=3):
    for _ in range(tries):
        txt = wait_text(driver, xpath, timeout=8)
        if txt: return txt
        time.sleep(delay)
    return None

# ======= 1 passo (trecho+ADVP) =======
def processar_trecho_advp(driver, base_tab, out_path, origin, destiny, advp, espera):
    data_voo = (datetime.now(TZ) + timedelta(days=advp)).date()
    url = build_url(origin, destiny, data_voo.strftime("%Y-%m-%d"))
    trecho_str = f"{origin}-{destiny}"

    try: driver.switch_to.window(base_tab)
    except Exception: base_tab = driver.current_window_handle
    navigate_same_tab(driver, url)

    status = wait_for_buy_button_or_no_flights(driver, espera)

    if status in ("no_flights", "timeout"):
        now = datetime.now(TZ)
        row = {
            "DATA DA BUSCA": now.date(),
            "HORA DA BUSCA": dtime(now.hour, now.minute, now.second),
            "TRECHO": trecho_str,
            "DATA PARTIDA": None,
            "HORA DA PARTIDA": None,
            "DATA CHEGADA": None,
            "HORA DA CHEGADA": None,
            "TARIFA": None,
            "TX DE EMBARQUE": None,
            "TOTAL": None,
            "CIA DO VOO": "Sem Ofertas",
        }
        append_row(out_path, row)
        return base_tab

    old_handles = set(driver.window_handles)
    if not js_click_first_buy(driver):
        now = datetime.now(TZ)
        row = {
            "DATA DA BUSCA": now.date(),
            "HORA DA BUSCA": dtime(now.hour, now.minute, now.second),
            "TRECHO": trecho_str,
            "DATA PARTIDA": None,
            "HORA DA PARTIDA": None,
            "DATA CHEGADA": None,
            "HORA DA CHEGADA": None,
            "TARIFA": None,
            "TX DE EMBARQUE": None,
            "TOTAL": None,
            "CIA DO VOO": "Sem Ofertas",
        }
        append_row(out_path, row)
        return base_tab

    switched = False
    try:
        WebDriverWait(driver, 6).until(lambda d: len(d.window_handles) > len(old_handles))
        new_handle = [h for h in driver.window_handles if h not in old_handles]
        if new_handle:
            driver.switch_to.window(new_handle[-1])
            switched = True
    except TimeoutException:
        pass

    # 2ª página
    partida_txt = wait_text_retry(driver, XPATH_PARTIDA, tries=3, delay=3)
    chegada_txt = wait_text_retry(driver, XPATH_CHEGADA, tries=3, delay=3)
    tarifa_txt  = wait_text_retry(driver, XPATH_TARIFA,  tries=3, delay=3)
    taxa_txt    = wait_text_retry(driver, XPATH_TAXA,    tries=3, delay=3)
    cia_txt     = wait_text_retry(driver, XPATH_CIA,     tries=2, delay=2)

    partida_dt = parse_datetime_br(partida_txt, fallback_date=data_voo)
    chegada_dt = parse_datetime_br(chegada_txt, fallback_date=data_voo)
    tarifa_val = brl_to_decimal(tarifa_txt)
    taxa_val   = brl_to_decimal(taxa_txt)
    total_val  = (tarifa_val or Decimal(0)) + (taxa_val or Decimal(0)) if (tarifa_val or taxa_val) else None

    data_partida = partida_dt.date() if partida_dt else None
    hora_partida = partida_dt.time() if partida_dt else None
    data_chegada = chegada_dt.date() if chegada_dt else None
    hora_chegada = chegada_dt.time() if chegada_dt else None

    cia_clean = clean_cia_text(cia_txt) or ("Sem Ofertas" if (tarifa_val is None and taxa_val is None) else "")

    now = datetime.now(TZ)
    row = {
        "DATA DA BUSCA": now.date(),
        "HORA DA BUSCA": dtime(now.hour, now.minute, now.second),
        "TRECHO": trecho_str,
        "DATA PARTIDA": data_partida,
        "HORA DA PARTIDA": hora_partida,
        "DATA CHEGADA": data_chegada,
        "HORA DA CHEGADA": hora_chegada,
        "TARIFA": float(tarifa_val) if tarifa_val is not None else None,
        "TX DE EMBARQUE": float(taxa_val) if taxa_val is not None else None,
        "TOTAL": float(total_val) if total_val is not None else None,
        "CIA DO VOO": cia_clean,
    }
    append_row(out_path, row)

    if switched:
        try: driver.close()
        except Exception: pass
        try: driver.switch_to.window(base_tab)
        except Exception:
            driver.switch_to.window(driver.window_handles[0])
            base_tab = driver.current_window_handle

    return base_tab

# ======= MAIN =======
def main():
    parser = argparse.ArgumentParser(description="FlipMilhas scraper p/ GitHub Actions (1 ciclo opcional).")
    parser.add_argument("--saida",    default="output", help="Pasta para salvar Excel")
    parser.add_argument("--file",     default="FLIPMILHAS.xlsx", help="Nome do arquivo Excel")
    parser.add_argument("--espera",   type=int, default=20, help="Segundos máx p/ 'Comprar' ou 'Nenhum voo'")
    parser.add_argument("--headless", action="store_true", help="Força headless")
    parser.add_argument("--gui",      dest="headless", action="store_false", help="Abre janela (debug local)")
    parser.add_argument("--once",     action="store_true", help="Roda apenas 1 ciclo e finaliza")
    parser.set_defaults(headless=True)  # no GitHub: headless por padrão
    args = parser.parse_args()

    os.makedirs(args.saida, exist_ok=True)
    out_path = os.path.join(args.saida, args.file)

    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s | %(levelname)s | %(message)s",
                        datefmt="%H:%M:%S")
    logging.info("Saída: %s | Planilha: %s | Aba: %s | Headless: %s", args.saida, args.file, SHEET_NAME, args.headless)

    driver, wait = setup_driver(headless=args.headless)
    base_tab = driver.current_window_handle

    try:
        def ciclo():
            nonlocal base_tab
            for (origin, destiny) in TRECHOS:
                for advp in ADVP_LIST:
                    try:
                        base_tab = processar_trecho_advp(
                            driver=driver, base_tab=base_tab, out_path=out_path,
                            origin=origin, destiny=destiny, advp=advp, espera=args.espera
                        )
                    except Exception as e:
                        logging.exception("Falha em %s-%s ADVP %d: %s", origin, destiny, advp, e)

        if args.once:
            ciclo()
        else:
            while True:
                ciclo()
                logging.info("Ciclo terminado. Aguardando 5 minutos…")
                time.sleep(300)

    finally:
        try: driver.quit()
        except Exception: pass
        logging.info("Finalizado.")

if __name__ == "__main__":
    main()
