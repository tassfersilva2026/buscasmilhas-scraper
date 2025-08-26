# capoviagens_scraper_gha.py — GitHub Actions / Local (1 ciclo por padrão)
# Colhe 1º card da Capo Viagens com estratégias anti-“skeleton”
# CLI: --saida data [--file CAPOVIAGENS.xlsx] --headless --once

import os, re, time, argparse, logging
from datetime import datetime, timedelta, date, time as dtime
from decimal import Decimal, InvalidOperation
from zoneinfo import ZoneInfo
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Selenium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

# ============================= CONFIG =============================
VERSION = "capo-gha-v2.0"
TZ = ZoneInfo("America/Sao_Paulo")
SHEET_NAME = "BUSCAS"
CENTER = Alignment(horizontal="center", vertical="center")

# Trechos/ADVP solicitados (os mesmos que você usa no local)
ADVP_LIST = [1, 3, 7, 14, 21, 30, 60, 90]
TRECHOS = [
    ("CGH", "SDU"), ("SDU", "CGH"),
    ("GRU", "POA"), ("POA", "GRU"),
    ("CGH", "GIG"), ("GIG", "CGH"),
    ("BSB", "CGH"), ("CGH", "BSB"),
    ("CGH", "REC"), ("REC", "CGH"),
    ("CGH", "SSA"), ("SSA", "CGH"),
    ("BSB", "GIG"), ("GIG", "BSB"),
    ("GIG", "REC"), ("REC", "GIG"),
    ("GIG", "SSA"), ("SSA", "GIG"),
    ("BSB", "SDU"), ("SDU", "BSB"),
]


# URL da Capo
def build_url(trecho: str, advp_days: int) -> tuple[str, date]:
    orig, dest = trecho.split("-")
    dep_date = (datetime.now(TZ).date() + timedelta(days=advp_days))
    url = ("https://www.capoviagens.com.br/voos/"
           f"?fromAirport={orig}&toAirport={dest}&departureDate={dep_date.strftime('%Y-%m-%d')}"
           "&adult=1&child=0&cabin=Basic&isTwoWays=false")
    return url, dep_date

# ============================= Regex/Heurísticas =============================
PRICE_RE = re.compile(r'R\$\s*\d{1,3}(?:\.\d{3})*,\d{2}')
TIME_RE  = re.compile(r'\b\d{2}:\d{2}(?::\d{2})?\b')

def wait_results_ready(driver, max_wait=25):
    """Espera até o <main> ter texto com 'R$' ou HH:MM (evita skeleton)."""
    t0 = time.time()
    while time.time() - t0 < max_wait:
        try:
            root = driver.find_element(By.XPATH, "//main")
            txt = (root.text or "").strip()
            if PRICE_RE.search(txt) or TIME_RE.search(txt):
                return True
        except Exception:
            pass
        time.sleep(1)
    return False

def scroll_jiggle(driver):
    try:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(0.3)
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(0.2)
    except Exception:
        pass

def first_visible_card(driver):
    """Tenta pegar o 1º card clicável da lista (robusto a pequenas mudanças)."""
    X_CANDIDATES = [
        "//main//label[contains(@class,'cursor-pointer')][1]",
        "//*[@id='__next']//main//label[1]",
        "//*[@id='__next']//main//div[@role='radiogroup']//label[1]",
    ]
    for xp in X_CANDIDATES:
        try:
            el = driver.find_element(By.XPATH, xp)
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            return el
        except Exception:
            continue
    return None

def text_or_inner(el, rel_xpath):
    """Pega .text do subelemento; se vier vazio, tenta innerText."""
    try:
        child = el.find_element(By.XPATH, rel_xpath)
        t = (child.text or "").strip()
        if not t:
            t = (child.get_attribute("innerText") or "").strip()
        return t or None
    except Exception:
        return None

def regex_from_inner_text(el):
    """Fallback: varre innerText do card e extrai horários/preços por regex."""
    out = {}
    try:
        txt = (el.get_attribute("innerText") or "").strip()
    except Exception:
        return out
    times = re.findall(r'\b\d{2}:\d{2}(?::\d{2})?\b', txt)
    prices_all = re.findall(r'R\$\s*\d{1,3}(?:\.\d{3})*,\d{2}', txt)
    if times:
        out["partida"] = times[0] + (":00" if len(times[0]) == 5 else "")
        if len(times) > 1:
            out["chegada"] = times[1] + (":00" if len(times[1]) == 5 else "")
    if prices_all:
        out["qualquer_preco"] = prices_all[0]
    return out

def brl_to_float(s):
    if not s: return None
    s1 = re.sub(r"[^\d,\.]", "", s).replace(".", "").replace(",", ".")
    try: return float(s1)
    except Exception: return None

def hhmm_to_time(s):
    if not s: return None
    m = re.fullmatch(r"(\d{2}):(\d{2})(?::(\d{2}))?", s.strip())
    if not m: return None
    hh, mm, ss = m.groups(); ss = ss or "00"
    try: return dtime(int(hh), int(mm), int(ss))
    except ValueError: return None

# ============================= Excel =============================
HEADERS = [
    "DATA DA BUSCA","HORA DA BUSCA","TRECHO",
    "DATA PARTIDA","HORA DA PARTIDA","DATA CHEGADA","HORA DA CHEGADA",
    "TARIFA","TX DE EMBARQUE","TOTAL","CIA DO VOO",
]

def _is_valid_xlsx(path: str) -> bool:
    try:
        if not os.path.isfile(path): return False
        with ZipFile(path, "r") as zf:
            return "[Content_Types].xml" in zf.namelist()
    except Exception:
        return False

def _create_new_workbook(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    widths = {
        "DATA DA BUSCA": 14, "HORA DA BUSCA": 12, "TRECHO": 12,
        "DATA PARTIDA": 14, "HORA DA PARTIDA": 14,
        "DATA CHEGADA": 14, "HORA DA CHEGADA": 14,
        "TARIFA": 14, "TX DE EMBARQUE": 16, "TOTAL": 14, "CIA DO VOO": 20,
    }
    for j, hdr in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(hdr, 16)
        ws.cell(row=1, column=j).alignment = CENTER
    wb.save(path); wb.close()

def ensure_workbook(path: str):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    if not os.path.exists(path) or not _is_valid_xlsx(path):
        _create_new_workbook(path); return
    wb = load_workbook(path)
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(HEADERS)
        for j in range(1, len(HEADERS)+1):
            ws.cell(row=1, column=j).alignment = CENTER
        wb.save(path)
    wb.close()

def append_row(path: str, row_values: dict):
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    if ws.max_row == 1 and all((c.value is None for c in ws[1])):
        ws.append(HEADERS)
        for j in range(1, len(HEADERS)+1):
            ws.cell(row=1, column=j).alignment = CENTER

    row = [row_values.get(h) for h in HEADERS]
    ws.append(row)
    r = ws.max_row

    fmt = {
        "DATA DA BUSCA": "DD/MM/YYYY",
        "HORA DA BUSCA": "HH:MM:SS",
        "DATA PARTIDA": "DD/MM/YYYY",
        "HORA DA PARTIDA": "HH:MM:SS",
        "DATA CHEGADA": "DD/MM/YYYY",
        "HORA DA CHEGADA": "HH:MM:SS",
        "TARIFA": "#,##0.00",
        "TX DE EMBARQUE": "#,##0.00",
        "TOTAL": "#,##0.00",
    }
    for c_idx, hdr in enumerate(HEADERS, start=1):
        cell = ws.cell(row=r, column=c_idx)
        if hdr in fmt and cell.value is not None:
            cell.number_format = fmt[hdr]
        cell.alignment = CENTER

    wb.save(path); wb.close()

# ============================= Selenium =============================
def _maybe_set_binary_location(opts: Options):
    chrome_path = os.getenv("CHROME_PATH") or os.getenv("GOOGLE_CHROME_SHIM")
    if chrome_path and os.path.exists(chrome_path):
        opts.binary_location = chrome_path

def setup_driver(headless=True):
    options = Options()
    if headless:
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--hide-scrollbars")
    _maybe_set_binary_location(options)
    # Flags estáveis no CI
    options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-sandbox")
    options.add_argument("--log-level=3")
    options.add_argument("--mute-audio")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--no-first-run")
    options.add_argument("--disable-notifications")
    options.add_argument("--lang=pt-BR")
    options.add_argument("--use-gl=swiftshader")
    options.add_argument("--enable-unsafe-swiftshader")

    service = ChromeService()
    driver = webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 15)
    return driver, wait

def navigate_same_tab(driver, url):
    try:
        driver.execute_script("window.location.assign(arguments[0]);", url)
    except Exception:
        driver.get(url)

# ============================= Core de 1 passo =============================
def processar_trecho_advp(driver, base_tab, out_path, trecho_str, advp, espera):
    now = datetime.now(TZ)
    data_voo = (now + timedelta(days=advp)).date()
    url, _dep = build_url(trecho_str, advp)

    try: driver.switch_to.window(base_tab)
    except Exception: base_tab = driver.current_window_handle
    navigate_same_tab(driver, url)

    # ======= BLOCO QUE VOCÊ ENVIOU (dentro da função!) =======
    # Espera conteúdo real
    scroll_jiggle(driver)
    ok = wait_results_ready(driver, max_wait=espera)
    now = datetime.now(TZ)

    if not ok:
        row = {
            "DATA DA BUSCA": now.date(),
            "HORA DA BUSCA": dtime(now.hour, now.minute, now.second),
            "TRECHO": trecho_str,
            "DATA PARTIDA": data_voo,
            "HORA DA PARTIDA": None,
            "DATA CHEGADA": data_voo,
            "HORA DA CHEGADA": None,
            "TARIFA": None,
            "TX DE EMBARQUE": None,
            "TOTAL": None,
            "CIA DO VOO": "Sem Ofertas",
        }
        append_row(out_path, row)
        return base_tab

    # 1º card visível
    card = first_visible_card(driver)
    if card is None:
        # fallback: usa o main inteiro por regex
        row = {
            "DATA DA BUSCA": now.date(),
            "HORA DA BUSCA": dtime(now.hour, now.minute, now.second),
            "TRECHO": trecho_str,
            "DATA PARTIDA": data_voo,
            "HORA DA PARTIDA": None,
            "DATA CHEGADA": data_voo,
            "HORA DA CHEGADA": None,
            "TARIFA": None,
            "TX DE EMBARQUE": None,
            "TOTAL": None,
            "CIA DO VOO": "Sem Ofertas",
        }
        append_row(out_path, row)
        return base_tab

    # tenta pelos XPaths originais RELATIVOS ao card
    rel_partida = ".//label/label//div/div/div[1]/span[1]"
    rel_chegada = ".//label/label//div/div/div[3]/span[1]"
    rel_tarifa  = ".//ancestor::div[1]/following::div[contains(@class,'price')][1]//span[1]"  # fallback amplo
    rel_tx_emb  = ".//ancestor::div[1]/following::div//span[contains(.,',')][last()]"
    rel_total   = ".//ancestor::div[1]/following::div//span[contains(.,'R$')][last()]"
    rel_cia     = ".//div[contains(@class,'div')]/span | .//div/span"

    # 1ª tentativa: XPaths “oficiais” (relativos ao card)
    hora_partida_txt = text_or_inner(card, rel_partida)
    hora_chegada_txt = text_or_inner(card, rel_chegada)
    tarifa_txt       = text_or_inner(card, rel_tarifa)
    tx_emb_txt       = text_or_inner(card, rel_tx_emb)
    total_txt        = text_or_inner(card, rel_total)
    cia_txt          = text_or_inner(card, rel_cia)

    # Fallback final: regex no innerText do card
    if not (hora_partida_txt and total_txt and (tarifa_txt or tx_emb_txt)):
        parsed = regex_from_inner_text(card)
        hora_partida_txt = hora_partida_txt or parsed.get("partida")
        hora_chegada_txt = hora_chegada_txt or parsed.get("chegada")
        if not (tarifa_txt or total_txt):
            anyp = parsed.get("qualquer_preco")
            tarifa_txt = tarifa_txt or anyp
            total_txt  = total_txt  or anyp

    # Normalizações
    hora_partida = hhmm_to_time(hora_partida_txt)
    hora_chegada = hhmm_to_time(hora_chegada_txt)
    tarifa_val   = brl_to_float(tarifa_txt)
    tx_emb_val   = brl_to_float(tx_emb_txt)
    total_val    = brl_to_float(total_txt)
    cia_clean    = (cia_txt or "").strip()

    row = {
        "DATA DA BUSCA": now.date(),
        "HORA DA BUSCA": dtime(now.hour, now.minute, now.second),
        "TRECHO": trecho_str,
        "DATA PARTIDA": data_voo,
        "HORA DA PARTIDA": hora_partida,
        "DATA CHEGADA": data_voo,
        "HORA DA CHEGADA": hora_chegada,
        "TARIFA": tarifa_val,
        "TX DE EMBARQUE": tx_emb_val,
        "TOTAL": total_val,
        "CIA DO VOO": cia_clean,
    }
    append_row(out_path, row)
    return base_tab
    # ======= FIM DO BLOCO =======

# ============================= MAIN =============================
def main():
    parser = argparse.ArgumentParser(description="Capo Viagens scraper (GH Actions / Local).")
    parser.add_argument("--saida",    default="data", help="Pasta para salvar Excel (default: data)")
    parser.add_argument("--file",     default="", help="Nome do arquivo .xlsx (se vazio, usa timestamp)")
    parser.add_argument("--espera",   type=int, default=25, help="Segundos p/ esperar render do 1º card")
    parser.add_argument("--headless", action="store_true", help="Headless")
    parser.add_argument("--gui",      dest="headless", action="store_false", help="Janela visível (debug)")
    parser.add_argument("--once",     action="store_true", help="Roda 1 ciclo e finaliza")
    parser.set_defaults(headless=True)
    args = parser.parse_args()

    os.makedirs(args.saida, exist_ok=True)
    # nome do arquivo
    if args.file:
        out_path = os.path.join(args.saida, args.file)
    else:
        stamp = datetime.now(TZ).strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(args.saida, f"CAPOVIAGENS_{stamp}.xlsx")

    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s | %(levelname)s | %(message)s",
                        datefmt="%H:%M:%S")
    logging.info("Versão: %s | Planilha: %s | Aba: %s | Headless: %s",
                 VERSION, out_path, SHEET_NAME, args.headless)

    ensure_workbook(out_path)
    driver, _wait = setup_driver(headless=args.headless)
    base_tab = driver.current_window_handle

    try:
        def ciclo():
            nonlocal base_tab
            for trecho in TRECHOS:
                for advp in ADVP_LIST:
                    try:
                        base_tab = processar_trecho_advp(
                            driver=driver,
                            base_tab=base_tab,
                            out_path=out_path,
                            trecho_str=trecho,
                            advp=advp,
                            espera=args.espera
                        )
                    except Exception as e:
                        logging.exception("Falha em %s ADVP %d: %s", trecho, advp, e)

        if args.once:
            ciclo()
        else:
            ciclo()  # no CI, geralmente usamos --once; aqui deixo 1 ciclo por padrão

    finally:
        try: driver.quit()
        except Exception: pass
        logging.info("Finalizado.")

# ============================= Driver/setup =============================
def setup_driver(headless=True):
    options = Options()
    if headless:
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--hide-scrollbars")
    _maybe_set_binary_location(options)
    options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-sandbox")
    options.add_argument("--log-level=3")
    options.add_argument("--mute-audio")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--no-first-run")
    options.add_argument("--disable-notifications")
    options.add_argument("--lang=pt-BR")
    options.add_argument("--use-gl=swiftshader")
    options.add_argument("--enable-unsafe-swiftshader")
    service = ChromeService()
    driver = webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 15)
    return driver, wait

if __name__ == "__main__":
    main()

